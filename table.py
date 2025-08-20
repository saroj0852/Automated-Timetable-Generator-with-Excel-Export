import pandas as pd
from ortools.sat.python import cp_model
# The openpyxl library is needed for advanced Excel formatting
try:
    from openpyxl import load_workbook
    from openpyxl.styles import Font, Alignment, Border, Side
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# --- 1. CONFIGURATION AND INPUT DATA ---
# This section defines all the core data for the university schedule.
DAYS = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday"]
ALL_SLOTS = ["9-10", "10-11", "11-12", "12-1", "2-3", "3-4", "4-5"]
LAB_SLOT_STARTS = ["9-10", "11-12", "3-4"] # Labs are 2 hours, so they can only start here
GROUPS = ['A', 'B']
SECTIONS = [
    'CSE-3-1', 'CSE-3-2', 'AIML-3',
    'CSE-5', 'IT-5',
    'CSE-7', 'IT-7'
]
SECTION_THEORY_ROOM = {
    'CSE-3-1': 'B-209', 'CSE-3-2': 'B-209', 'AIML-3': 'A-32',
    'CSE-5': 'B-205', 'IT-5': 'B-205', 'CSE-7': 'D-303', 'IT-7': 'D-303'
}
LAB_ROOMS = ['CS105', 'CS106', 'CS107', 'CS115', 'CS204', 'CS205', 'CS208', 'CS220']
SUBJECTS = {
    'CSE-3-1': [('DLD', 'AM'), ('DS', 'SK'), ('DBE', 'SP'), ('OOP', 'AVL')],
    'CSE-3-2': [('DLD', 'SWP'), ('DS', 'GS'), ('DBE', 'KN'), ('OOP', 'SS')],
    'AIML-3':  [('DLD', 'PKS'), ('DS', 'SKS'), ('DBE', 'GF4'), ('OOP', 'SKB')],
    'CSE-5':   [('TOC', 'GF2'), ('OS', 'HSB'), ('AI/ML', 'PKD'), ('CNS', 'SPanda')],
    'IT-5':    [('TOC', 'KKS'), ('OS', 'SBD'), ('AI/ML', 'GF3'), ('DMDW', 'BN')],
    'CSE-7':   [('AI', 'GF5'), ('IWP', 'GF6'), ('IP', 'KN')],
    'IT-7':    [('AI', 'MRS'), ('CS', 'AD'), ('SM', 'SKN')]
}
LABS = {
    'CSE-3-1': ['DLD Lab', 'DS Lab', 'DBE Lab', 'OOP Lab'],
    'CSE-3-2': ['DLD Lab', 'DS Lab', 'DBE Lab', 'OOP Lab'],
    'AIML-3':  ['DLD Lab', 'DS Lab', 'DBE Lab', 'OOP Lab'],
    'CSE-5':   ['TOC Lab', 'OS Lab', 'AI/ML Lab'],
    'IT-5':    ['TOC Lab', 'OS Lab', 'AI/ML Lab'],
    'CSE-7':   ['Web Programming Lab'],
    'IT-7':    ['Artificial Intelligence Lab']
}

# Helper function to find the teacher for a given lab based on the theory subject
def get_teacher_for_lab(section, lab_name):
    # Maps specific lab names to their corresponding theory subject codes
    subject_name_map = {'Web Programming Lab': 'IWP', 'Artificial Intelligence Lab': 'AI'}
    # Derives the subject name from the lab name, using the map or by removing " Lab"
    subject_name = subject_name_map.get(lab_name, lab_name.replace(' Lab', ''))
    # Finds the teacher assigned to that subject for the given section
    for subject, teacher in SUBJECTS.get(section, []):
        if subject == subject_name:
            return teacher
    return None

# Create comprehensive lists of all teachers and rooms for easier constraint definition
ALL_TEACHERS = sorted(list(set(teacher for section_subjects in SUBJECTS.values() for _, teacher in section_subjects)))
ALL_ROOMS = sorted(list(set(SECTION_THEORY_ROOM.values()) | set(LAB_ROOMS)))

# --- 2. THE SCHEDULING MODEL ---
model = cp_model.CpModel()

# Create a dictionary to hold all the boolean variables.
# Each variable represents whether a specific class (theory or lab) is scheduled at a specific time and place.
class_vars = {}

# Create variables for all possible theory classes
for section in SECTIONS:
    theory_room = SECTION_THEORY_ROOM.get(section)
    if theory_room:
        for subject, teacher in SUBJECTS.get(section, []):
            for day_idx in range(len(DAYS)):
                for slot_idx in range(len(ALL_SLOTS)):
                    var_name = f'theory_{section}_{subject}_{day_idx}_{slot_idx}'
                    # Key: (section, group, subject, teacher, day, slot, room)
                    class_vars[(section, 'ALL', subject, teacher, day_idx, slot_idx, theory_room)] = model.NewBoolVar(var_name)

# Create variables for all possible lab classes
for section in LABS:
    for lab_name in LABS.get(section, []):
        teacher = get_teacher_for_lab(section, lab_name)
        if not teacher: continue
        for group in GROUPS:
            for day_idx in range(len(DAYS)):
                for slot_idx, slot in enumerate(ALL_SLOTS):
                    # Labs can only start at specific times
                    if slot in LAB_SLOT_STARTS:
                        for room in LAB_ROOMS:
                            var_name = f'lab_{section}_{group}_{lab_name}_{day_idx}_{slot_idx}_{room}'
                            class_vars[(section, group, lab_name, teacher, day_idx, slot_idx, room)] = model.NewBoolVar(var_name)

# --- 3. HARD CONSTRAINTS ---
# These are rules that absolutely cannot be broken.

# Rule: A room can only have one class at a time.
for day_idx in range(len(DAYS)):
    for slot_idx in range(len(ALL_SLOTS)):
        for room in ALL_ROOMS:
            active_classes_in_slot = []
            for (sec, grp, subj, tc, d, s, rm), var in class_vars.items():
                if d == day_idx and rm == room:
                    # Check if the class is active in this slot.
                    # A lab is active in its starting slot and the one immediately after.
                    is_active_in_slot = (s == slot_idx) or ('Lab' in subj and s == slot_idx - 1)
                    if is_active_in_slot:
                        active_classes_in_slot.append(var)
            model.AddAtMostOne(active_classes_in_slot)

# Rule: A teacher can only teach one class at a time.
for day_idx in range(len(DAYS)):
    for slot_idx in range(len(ALL_SLOTS)):
        for teacher in ALL_TEACHERS:
            active_classes_for_teacher = []
            for (sec, grp, subj, tc, d, s, rm), var in class_vars.items():
                if d == day_idx and tc == teacher:
                    is_active_in_slot = (s == slot_idx) or ('Lab' in subj and s == slot_idx - 1)
                    if is_active_in_slot:
                        active_classes_for_teacher.append(var)
            model.AddAtMostOne(active_classes_for_teacher)

# Rule: A section (or a group within a section) can only have one class at a time.
for section in SECTIONS:
    # This loop covers the whole section ('ALL') and individual groups ('A', 'B')
    for group in GROUPS + ['ALL']:
        for day_idx in range(len(DAYS)):
            for slot_idx in range(len(ALL_SLOTS)):
                active_classes_for_section_group = []
                for (sec, grp, subj, tc, d, s, rm), var in class_vars.items():
                    if sec == section and d == day_idx:
                        # Check if the class is for the whole section or a specific group
                        is_relevant_group = (grp == group) or (group == 'ALL' and grp in GROUPS)
                        if is_relevant_group:
                            is_active_in_slot = (s == slot_idx) or ('Lab' in subj and s == slot_idx - 1)
                            if is_active_in_slot:
                                active_classes_for_section_group.append(var)
                model.AddAtMostOne(active_classes_for_section_group)

# Rule: Each theory subject must be taught exactly 3 times a week for each section.
for section, section_subjects in SUBJECTS.items():
    for subject, teacher in section_subjects:
        model.Add(sum(var for (sec, grp, subj, tc, d, s, rm), var in class_vars.items()
                      if sec == section and subj == subject) == 3)

# Rule: Each lab for each group must be scheduled exactly once a week.
for section, section_labs in LABS.items():
    for lab_name in section_labs:
        for group in GROUPS:
            model.Add(sum(var for (sec, grp, subj, tc, d, s, rm), var in class_vars.items()
                          if sec == section and grp == group and subj == lab_name) == 1)

# Rule: A teacher cannot teach in the 2-3 slot if they taught in the 12-1 slot on the same day (recess rule).
slot_12_1_idx, slot_2_3_idx = 3, 4
for teacher in ALL_TEACHERS:
    for day_idx in range(len(DAYS)):
        # Variables to check if teacher is busy at these specific times
        busy_at_12 = model.NewBoolVar(f'busy_{teacher}_{day_idx}_at12')
        busy_at_2 = model.NewBoolVar(f'busy_{teacher}_{day_idx}_at2')
        
        # Link busy_at_12 to whether the teacher has any class (theory or lab) in the 12-1 slot
        classes_at_12 = [var for (s,g,sb,t,d,sl,r), var in class_vars.items() if t == teacher and d == day_idx and ((sl == slot_12_1_idx) or ('Lab' in sb and sl == slot_12_1_idx - 1))]
        model.Add(sum(classes_at_12) > 0).OnlyEnforceIf(busy_at_12)
        model.Add(sum(classes_at_12) == 0).OnlyEnforceIf(busy_at_12.Not())

        # Link busy_at_2 to whether the teacher has any class in the 2-3 slot
        classes_at_2 = [var for (s,g,sb,t,d,sl,r), var in class_vars.items() if t == teacher and d == day_idx and sl == slot_2_3_idx]
        model.Add(sum(classes_at_2) > 0).OnlyEnforceIf(busy_at_2)
        model.Add(sum(classes_at_2) == 0).OnlyEnforceIf(busy_at_2.Not())
        
        # The actual constraint: It's not possible to be busy at both times.
        model.AddBoolOr([busy_at_12.Not(), busy_at_2.Not()])

# Rule: Each section can have a maximum of 4 theory classes per day.
for section in SECTIONS:
    for day_idx in range(len(DAYS)):
        daily_theory_classes = [
            var for (sec, grp, subj, tc, d, s, rm), var in class_vars.items()
            if sec == section and d == day_idx and 'Lab' not in subj
        ]
        model.Add(sum(daily_theory_classes) <= 4)

# Rule: When two groups of a section have labs simultaneously, they must be for different subjects.
for section in SECTIONS:
    section_labs = LABS.get(section, [])
    if not section_labs: continue
    for day_idx in range(len(DAYS)):
        for slot_idx, slot in enumerate(ALL_SLOTS):
            if slot in LAB_SLOT_STARTS:
                for lab_name in section_labs:
                    # Sum of variables for Group A doing a specific lab at this time (will be 0 or 1).
                    group_a_specific_lab = [
                        var for (sec, grp, subj, tc, d, s, rm), var in class_vars.items()
                        if sec == section and grp == 'A' and subj == lab_name and d == day_idx and s == slot_idx
                    ]
                    # Sum of variables for Group B doing the same lab at the same time.
                    group_b_specific_lab = [
                        var for (sec, grp, subj, tc, d, s, rm), var in class_vars.items()
                        if sec == section and grp == 'B' and subj == lab_name and d == day_idx and s == slot_idx
                    ]
                    # The sum must be at most 1, meaning both groups cannot be scheduled for the same lab at the same time.
                    model.Add(sum(group_a_specific_lab) + sum(group_b_specific_lab) <= 1)


# --- 4. SOFT CONSTRAINTS AND OBJECTIVE FUNCTION ---
# These are rules that the solver should try to follow but can break if necessary.
# We turn them into penalties and ask the solver to minimize the total penalty.

# Goal 1: Balance teacher workload across the week.
max_daily_load_vars = []
for teacher in ALL_TEACHERS:
    # This variable will hold the maximum number of classes this teacher has on any day.
    max_load_for_teacher = model.NewIntVar(0, len(ALL_SLOTS), f'max_load_{teacher}')
    daily_loads = []
    for day_idx in range(len(DAYS)):
        # Sum up all classes for the teacher on this day.
        classes_on_day = sum(
            var for (sec, grp, subj, tc, d, s, rm), var in class_vars.items()
            if tc == teacher and d == day_idx
        )
        daily_loads.append(classes_on_day)
    # Tell the model that max_load_for_teacher must be the maximum of the daily loads.
    model.AddMaxEquality(max_load_for_teacher, daily_loads)
    max_daily_load_vars.append(max_load_for_teacher)

# Goal 2: Encourage theory classes to be in continuous blocks.
continuity_penalties = []
for section in SECTIONS:
    for day_idx in range(len(DAYS)):
        # Create helper variables to see if a theory class is in a given slot
        is_theory_scheduled = {}
        for slot_idx in range(len(ALL_SLOTS)):
            is_theory_scheduled[slot_idx] = model.NewBoolVar(f'is_theory_{section}_{day_idx}_{slot_idx}')
            theory_in_slot = [
                var for (sec, grp, subj, tc, d, s, rm), var in class_vars.items()
                if sec == section and d == day_idx and s == slot_idx and 'Lab' not in subj
            ]
            # Link the helper variable to the actual class variables
            model.Add(sum(theory_in_slot) > 0).OnlyEnforceIf(is_theory_scheduled[slot_idx])
            model.Add(sum(theory_in_slot) == 0).OnlyEnforceIf(is_theory_scheduled[slot_idx].Not())

        # Create penalty variables for each potential transition
        for i in range(len(ALL_SLOTS) - 1):
            transition_var = model.NewBoolVar(f'transition_{section}_{day_idx}_{i}')
            # A transition occurs if a slot is filled and the next is empty, or vice-versa.
            model.Add(is_theory_scheduled[i] != is_theory_scheduled[i+1]).OnlyEnforceIf(transition_var)
            model.Add(is_theory_scheduled[i] == is_theory_scheduled[i+1]).OnlyEnforceIf(transition_var.Not())
            continuity_penalties.append(transition_var)

# Goal 3: Prefer, but do not require, that labs for Group A and B run in parallel.
parallel_lab_penalties = []
for section in SECTIONS:
    if section in LABS and len(LABS[section]) > 1:
        for day_idx in range(len(DAYS)):
            for slot_idx, slot in enumerate(ALL_SLOTS):
                if slot in LAB_SLOT_STARTS:
                    # Helper variable: True if Group A has any lab at this time
                    group_a_has_lab = model.NewBoolVar(f'gA_has_lab_{section}_{day_idx}_{slot_idx}')
                    # Helper variable: True if Group B has any lab at this time
                    group_b_has_lab = model.NewBoolVar(f'gB_has_lab_{section}_{day_idx}_{slot_idx}')

                    group_a_labs_at_time = [var for (sec, grp, subj, tc, d, s, rm), var in class_vars.items() if sec == section and grp == 'A' and d == day_idx and s == slot_idx and 'Lab' in subj]
                    group_b_labs_at_time = [var for (sec, grp, subj, tc, d, s, rm), var in class_vars.items() if sec == section and grp == 'B' and d == day_idx and s == slot_idx and 'Lab' in subj]

                    # Link helper variables to the actual sum of labs
                    model.Add(sum(group_a_labs_at_time) == 1).OnlyEnforceIf(group_a_has_lab)
                    model.Add(sum(group_a_labs_at_time) == 0).OnlyEnforceIf(group_a_has_lab.Not())
                    model.Add(sum(group_b_labs_at_time) == 1).OnlyEnforceIf(group_b_has_lab)
                    model.Add(sum(group_b_labs_at_time) == 0).OnlyEnforceIf(group_b_has_lab.Not())

                    # A penalty is incurred if one group has a lab but the other doesn't.
                    is_unbalanced = model.NewBoolVar(f'unbalanced_lab_{section}_{day_idx}_{slot_idx}')
                    model.Add(group_a_has_lab != group_b_has_lab).OnlyEnforceIf(is_unbalanced)
                    model.Add(group_a_has_lab == group_b_has_lab).OnlyEnforceIf(is_unbalanced.Not())
                    parallel_lab_penalties.append(is_unbalanced)

# Goal 4 (FIX): Prefer, but do not require, only one lab session per day per section.
daily_lab_session_penalties = []
for section in SECTIONS:
    for day_idx in range(len(DAYS)):
        daily_lab_sessions = []
        for slot_idx, slot in enumerate(ALL_SLOTS):
            if slot in LAB_SLOT_STARTS:
                is_lab_session_at_slot = model.NewBoolVar(f'lab_session_{section}_{day_idx}_{slot_idx}')
                labs_at_this_time = [var for (sec, grp, subj, tc, d, s, rm), var in class_vars.items() if sec == section and d == day_idx and s == slot_idx and 'Lab' in subj]
                if not labs_at_this_time:
                    model.Add(is_lab_session_at_slot == 0)
                    continue
                model.Add(sum(labs_at_this_time) > 0).OnlyEnforceIf(is_lab_session_at_slot)
                model.Add(sum(labs_at_this_time) == 0).OnlyEnforceIf(is_lab_session_at_slot.Not())
                daily_lab_sessions.append(is_lab_session_at_slot)
        
        # Penalize if the number of lab sessions in a day is more than 1.
        num_daily_sessions = sum(daily_lab_sessions)
        # Create a penalty variable that is max(0, num_daily_sessions - 1)
        penalty_var = model.NewIntVar(0, len(LAB_SLOT_STARTS) - 1, f'daily_lab_penalty_{section}_{day_idx}')
        model.Add(penalty_var >= num_daily_sessions - 1)
        daily_lab_session_penalties.append(penalty_var)


# Define the final objective: Minimize a weighted sum of all penalties.
WORKLOAD_PENALTY_WEIGHT = 10      # Strongly prioritize balancing teacher workload
PARALLEL_LAB_PENALTY_WEIGHT = 5   # Also strongly prefer parallel labs
DAILY_LAB_PENALTY_WEIGHT = 3      # Moderately prefer one lab session per day
CONTINUITY_PENALTY_WEIGHT = 1     # Weakly prefer continuous theory blocks

total_workload_penalty = sum(max_daily_load_vars)
total_continuity_penalty = sum(continuity_penalties)
total_parallel_lab_penalty = sum(parallel_lab_penalties)
total_daily_lab_penalty = sum(daily_lab_session_penalties)

model.Minimize(
    (WORKLOAD_PENALTY_WEIGHT * total_workload_penalty) +
    (CONTINUITY_PENALTY_WEIGHT * total_continuity_penalty) +
    (PARALLEL_LAB_PENALTY_WEIGHT * total_parallel_lab_penalty) +
    (DAILY_LAB_PENALTY_WEIGHT * total_daily_lab_penalty)
)


# --- 5. SOLVE AND EXPORT TO EXCEL ---
solver = cp_model.CpSolver()
# Set a time limit for the solver (e.g., 120 seconds)
solver.parameters.max_time_in_seconds = 120.0
status = solver.Solve(model)

if status == cp_model.FEASIBLE or status == cp_model.OPTIMAL:
    print("✅ Solution Found! Exporting Master Timetables to Excel...")
    
    excel_path = 'University_Master_Timetable.xlsx'
    with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
        # Create a separate master timetable sheet for each day
        for day_idx, day in enumerate(DAYS):
            # Initialize an empty DataFrame with sections as rows and slots as columns
            master_df = pd.DataFrame(index=SECTIONS, columns=ALL_SLOTS)
            master_df = master_df.fillna('') # Fill NaN with empty strings

            # Populate the DataFrame with the solved schedule for the current day
            for (section, group, subj, teacher, d_idx, slot_idx, room), var in class_vars.items():
                if d_idx == day_idx and solver.Value(var):
                    # Format the entry for the timetable cell
                    entry_group = f"({group})" if group != 'ALL' else ""
                    entry = f"{subj} {entry_group}\n({teacher})\n{room}"
                    
                    # Add entry to the cell. If a cell is already occupied (e.g. by two lab groups), append the new entry.
                    current_cell_value = master_df.at[section, ALL_SLOTS[slot_idx]]
                    if current_cell_value:
                        master_df.at[section, ALL_SLOTS[slot_idx]] += f"\n---\n{entry}"
                    else:
                        master_df.at[section, ALL_SLOTS[slot_idx]] = entry

                    # If it's a lab, it spans two slots. Fill the next cell as well.
                    if 'Lab' in subj and slot_idx + 1 < len(ALL_SLOTS):
                        next_slot = ALL_SLOTS[slot_idx + 1]
                        next_cell_value = master_df.at[section, next_slot]
                        if next_cell_value:
                             master_df.at[section, next_slot] += f"\n---\n{entry}"
                        else:
                             master_df.at[section, next_slot] = entry
            
            # Write the completed DataFrame for the day to a new Excel sheet
            master_df.to_excel(writer, sheet_name=day)

    print(f"✅ Master Timetables successfully saved to '{excel_path}'")

    # --- Optional: Apply advanced formatting to the Excel file ---
    if OPENPYXL_AVAILABLE:
        try:
            wb = load_workbook(excel_path)
            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            for ws in wb.worksheets:
                ws.column_dimensions['A'].width = 15 # Widen the first column (Sections)
                # Set width for time slot columns
                for col in ws.iter_cols(min_col=2, max_col=ws.max_column):
                    ws.column_dimensions[col[0].column_letter].width = 25
                # Set height for all rows to accommodate multi-line text
                for row in ws.iter_rows():
                    ws.row_dimensions[row[0].row].height = 60
                    for cell in row:
                        # Apply center alignment, text wrapping, and borders to all cells
                        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        cell.border = thin_border
                        # Make headers bold
                        if cell.row == 1 or cell.column == 1:
                            cell.font = Font(bold=True)
            wb.save(excel_path)
            print("✅ Excel file formatted successfully.")
        except Exception as e:
            print(f"Could not apply formatting to Excel file. Error: {e}")

else:
    print("❌ No solution found.")
    print("This could be because the constraints are too tight (e.g., not enough rooms or available teacher slots).")
    print("Try relaxing some constraints or increasing the solver's time limit.")
    print("Solver status:", solver.StatusName(status))
